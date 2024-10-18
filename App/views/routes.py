import json, requests, secrets, os, random, string, markdown2, google.generativeai as genai, smtplib

from flask import Blueprint, current_app, request, render_template, flash, redirect, url_for, send_from_directory, jsonify
from flask_admin.base import expose, AdminIndexView, Admin
from flask_login import login_required, current_user
from flask_mail import Message
from flask_ckeditor.utils import cleanify
from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import asc, desc
from datetime import datetime, timedelta, date
from babel.numbers import format_currency
from werkzeug.utils import secure_filename
from functools import wraps
from textwrap import shorten
from dotenv import load_dotenv

from App.models import User, DataPangan, Forum, Kebun, Artikel
from App import db, flatpages, mail

load_dotenv()
genai.configure(api_key=os.environ["GEMINI_API_KEY"])

views = Blueprint('views', __name__)

PICTURE_ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
REPORT_ALLOWED_EXTENSIONS = {'xlsx'}
REPORT_STAT = {'panen', 'penanaman'}

def picture_allowed_file(filename):
    return '.' in filename and \
            filename.rsplit('.', 1)[1].lower() in PICTURE_ALLOWED_EXTENSIONS

# Fungsi yang disederhanakan untuk memeriksa nama file dan tipe impor
def allowed_report_stat(filename, import_type):
    file_name_without_extension = os.path.splitext(filename)[0].lower()
    return import_type in file_name_without_extension

def report_allowed_file(filename):
    return '.' in filename and \
            filename.rsplit('.', 1)[1].lower() in REPORT_ALLOWED_EXTENSIONS

def generate_unique_id(prefix="KR_", string_length=2, number_length=4):
    """
    Generates a unique ID in the format KR_AB1234.

    Args:
        prefix: The static identifier prefix (default: "KR_").
        string_length: The length of the random string part (default: 2).
        number_length: The length of the random number part (default: 4).

    Returns:
        A unique ID string.
    """
    random_string = ''.join(random.choices(string.ascii_uppercase, k=string_length))
    random_number = ''.join(random.choices(string.digits, k=number_length))
    unique_id = f"{prefix}{random_string}{random_number}"
    return unique_id

def generate_username(email):
    """Generate a username from the email address."""
    username_base = email.split('@')[0]
    random_digits = ''.join(random.choice(string.digits) for _ in range(4))
    return f"{username_base}{random_digits}"

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role != role or not current_user.is_verified:
                flash('Anda tidak memiliki izin untuk mengakses halaman ini, silakan upgrade akun anda terlebih dahulu sebagai PETANI', 'warning')
                return redirect(url_for('views.index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Constants for API URL parameters
KAB_KOTA = 458  # Ternate
KOMODITAS_ID = 3
TARGET_KOMODITAS = ["Cabai Merah Keriting", "Cabai Rawit Merah", "Bawang Merah"]

# Helper function to fetch and format price data from API
def fetch_price_data(start_date, end_date):
    """Fetches price data from the API and formats it for display.

    Args:
        start_date (str): The starting date in YYYY-MM-DD format.
        end_date (str): The ending date in YYYY-MM-DD format.

    Returns:
        list: A list of dictionaries containing formatted price data.
    """

    url = f"https://panelharga.badanpangan.go.id/data/kabkota-range-by-levelharga/{KAB_KOTA}/{KOMODITAS_ID}/{start_date}/{end_date}"

    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for HTTP errors

        data = response.json()

        table_data = []
        for item in data["data"]:
            if item["name"] in TARGET_KOMODITAS:
                for date_data in item["by_date"]:
                    date_obj = datetime.strptime(date_data["date"], "%Y-%m-%d")
                    formatted_date = date_obj.strftime("%d/%m/%Y")
                    geomean_value = date_data["geomean"]

                    # Menyederhanakan format harga
                    formatted_price = "-" if geomean_value == "-" else format_currency(float(geomean_value), "IDR", locale="id_ID", decimal_quantization=False)[:-3]

                    table_data.append({
                        "date": formatted_date,
                        "name": item["name"],
                        "price": formatted_price
                    })

        return table_data

    except requests.exceptions.RequestException as e:
        flash(f"Error fetching data: {e}", category='error')
        return []  # Return an empty list on error
    
@views.route('/api/gemini', methods=['POST'])
def gemini_api():
    user_message = request.json.get('message')
    if not user_message:
        return jsonify({'error': 'Message is required'}), 400

    prompt = f"Saya adalah asisten virtual untuk platform agrikultur digital bernama RINDANG, yang membantu petani mengelola produksi pertanian dan memberikan informasi seputar pertanian di Kota Ternate. Saya hanya boleh memberikan jawaban terkait agrikultur, termasuk tetapi tidak terbatas pada: cara merawat tanaman, rekomendasi pupuk, langkah-langkah menghadapi cuaca, dan teknologi pertanian. Pertanyaan pengguna: {user_message}."

    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt)
        
        if response and response.text:
            # Convert Markdown to HTML in the backend using markdown2
            assistant_reply = markdown2.markdown(response.text)
        else:
            return jsonify({'error': 'No content received from Gemini API'}), 500

        return jsonify({'reply': assistant_reply}), 200

    except Exception as e:
        print(f"Error communicating with Gemini API: {e}")
        return jsonify({'error': f'Error communicating with Gemini API: {e}'}), 500

@views.route('/virtual-assistant')
def virtual_assistant():
    return render_template('features/virtual_assistant.html')

@views.route('/', methods=['GET'])
def index():
    kebun = Kebun.query.all()
    produksi = DataPangan.query.all()

    jml_kebun = len(kebun)
    total_panen = sum(prod.jml_panen for prod in produksi)

    # Calculate date range for the API
    today = datetime.today()
    one_week_ago = today - timedelta(days=7)
    start_date = one_week_ago.strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    return render_template('index.html', 
                            kebun=jml_kebun,
                            round=round, 
                            produksi=total_panen, 
                            start_date=start_date, 
                            end_date=end_date)

@views.route('/api/get-price-data', methods=['GET'])
def getpricedata():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        today = datetime.today()
        one_week_ago = today - timedelta(days=7)
        start_date = one_week_ago.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

    table_data = fetch_price_data(start_date, end_date)
    return jsonify(table_data)

@views.route('api/price-data', methods=['GET', 'POST'])
def get_price_data():
    kab_kota = request.args.get('kab_kota')
    komoditas_id = request.args.get('komoditas_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    url = f"https://panelharga.badanpangan.go.id/data/kabkota-range-by-levelharga/{KAB_KOTA}/{KOMODITAS_ID}/{start_date}/{end_date}"
    
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an error for bad responses
        return jsonify(response.json()), 200
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 500

@views.route('/personal', methods=['GET', 'POST'])
@login_required
def personal():
    users = User.query.filter_by(id=current_user.id)

    pagination_pages = 5
    page = request.args.get('page', 1, type=int)
    articles_pagination = Artikel.query.paginate(page=page, per_page=pagination_pages) #<-- MOVE THIS LINE OUT OF IF BLOCK!
    forum_pagination = Forum.query.paginate(page=page, per_page=pagination_pages) #<-- MOVE THIS LINE OUT OF IF BLOCK!

    if not articles_pagination.items: #Check if articles exist on the CURRENT page
        articles = []  # If not any article found for the page assign an empty array and handle no data condition at template
    else:
        articles = articles_pagination.items

    if not forum_pagination.items: #Check if articles exist on the CURRENT page
        forum = []  # If not any article found for the page assign an empty array and handle no data condition at template
    else:
        forum = forum_pagination.items

    return render_template('personal/index.html', users=users, articles=articles, forum=forum, min=min, max=max, articles_pagination=articles_pagination, forum_pagination=forum_pagination)

@views.route('/personal/profile')
@login_required
def personalprofile():
    if current_user.role == 'petani':
        return redirect(url_for('views.profil'))

    return render_template('personal/profile.html')

@views.route('/read_article/<int:id>', methods=['GET'])
def read_article(id):
    article = Artikel.query.get_or_404(id)
    if article.is_drafted:
        flash('Artikel tidak tersedia!', 'warning')
        return redirect(request.referrer)
    created_by = User.query.filter_by(id=article.created_by).first()
    return render_template('features/article.html', article=article, created_by=created_by.nama_lengkap)

@views.route('/write_article', methods=['GET', 'POST'])
@login_required
def write_article():
    articles = Artikel.query.filter_by(created_by=current_user.id)

    if request.method == 'POST':
        judul = request.form['judul']
        # konten = request.form['konten']
        konten = cleanify(request.form.get('ckeditor'))
        if request.form['action'] == 'posting':
            add_article = Artikel(judul=judul, content=konten, created_by=current_user.id)
        elif request.form['action'] == 'save':
            add_article = Artikel(judul=judul, content=konten, created_by=current_user.id, is_drafted=True)
        db.session.add(add_article)
        db.session.commit()
        flash('Berhasil membuat artikel', 'success')
        return redirect(url_for('views.personal'))
    return render_template('features/write_article.html', articles=articles)

@views.route('/update_article/<int:id>', methods=['GET', 'POST'])
@login_required
def update_article(id):
    article = Artikel.query.filter_by(created_by=current_user.id, id=id).first()

    if request.method == 'POST':
        article.judul = request.form['judul']
        article.content = request.form.get('ckeditor')
        if request.form['action'] == 'posting':
            article.is_drafted = False
        elif request.form['action'] == 'save':
            article.is_drafted = True
        db.session.commit()
        flash('Berhasil memperbarui artikel', 'success')
        return redirect(url_for('views.personal'))
    return render_template('features/update_article.html', article=article)

@views.route('delete_article/<int:id>', methods=['GET', 'POST'])
@login_required
def delete_article(id):
    article = Artikel.query.filter_by(created_by=current_user.id, id=id).first()
    db.session.delete(article)
    db.session.commit()
    flash('Berhasil menghapus artikel', 'warning')
    return redirect(url_for('views.personal'))

@views.route('/dashboard', methods=['GET', 'POST'])
@login_required
@role_required('petani')
def dashboard():
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))
    if current_user.role == 'admin':
        return redirect(url_for('admin_page.index'))

    data_pangan = DataPangan.query.filter_by(user_id=current_user.id).all()
    total_panen = sum(prod.jml_panen for prod in data_pangan if data_pangan)

    today = date.today()
    harvest_data = []
    next_harvest_days = None

    for estPanen in data_pangan:
        if estPanen.estimasi_panen:
            est_date = datetime.strptime(estPanen.estimasi_panen, '%Y-%m-%d').date()
            days_remaining = (est_date - today).days
            harvest_data.append({
                'date': estPanen.estimasi_panen,
                'days_remaining': days_remaining
            })

            # Update next_harvest_days if this is the soonest upcoming harvest
            if days_remaining > 0 and (next_harvest_days is None or days_remaining < next_harvest_days):
                next_harvest_days = days_remaining

    return render_template('dashboard/index.html', 
                            total_panen=total_panen, 
                            round=round, 
                            harvest_data=json.dumps(harvest_data),
                            next_harvest_days=next_harvest_days
                            )

@views.route('/dashboard/penjualan')
@login_required
@role_required('petani')
def penjualan():
    return render_template('dashboard/penjualan.html')

@views.route('/dashboard/harga-pangan', methods=['POST', 'GET'])
@login_required
# @role_required('petani')
def hargapangan():
    if current_user.role == 'admin':
        return redirect(url_for('admin_page.index'))

    # Menggunakan metode today() untuk mendapatkan tanggal hari ini
    # today = datetime.today()
    # one_week_ago = today - timedelta(days=7)
    # start_date = one_week_ago.strftime("%Y-%m-%d")
    # end_date = today.strftime("%Y-%m-%d")

    # table_data = fetch_price_data(start_date, end_date)

    return render_template('dashboard/harga-pangan.html')

@views.route('/harga-komoditas', methods=['GET', 'POST'])
def hargakomoditas():
    return render_template('features/harga_komoditas.html')

@views.route('/dashboard/data-pangan', methods=['POST','GET'])
@login_required
@role_required('petani')
def dataproduksi():
    if current_user.role == 'admin':
        return redirect(url_for('admin_page.index'))

    user_data = User.query.filter_by(id=current_user.id).first()
    pangan = DataPangan.query.filter_by(user_id=current_user.id).all()
    kebun = Kebun.query.filter_by(user_id=current_user.id).all()

    page = request.args.get('page', 1, type=int)
    per_page = 5

    sort_by = request.args.get('sort_by', 'tanggal_panen')
    sort_order = request.args.get('sort_order', 'desc')

    # Define a dictionary to map sort_by values to column names
    sort_columns = {
        'kebun': DataPangan.kebun_id,
        'bibit': DataPangan.jml_bibit,
        'tanam': DataPangan.tanggal_bibit,
        'status': DataPangan.status,
        'hasil': DataPangan.jml_panen
    }

    # Get the column to sort by
    sort_column = sort_columns.get(sort_by, DataPangan.tanggal_panen)

    # Apply the sorting
    if sort_order == 'asc':
        cabai = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Cabai').order_by(sort_column.asc())
        tomat = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Tomat').order_by(sort_column.asc())
    else:
        cabai = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Cabai').order_by(sort_column.desc())
        tomat = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Tomat').order_by(sort_column.desc())

    cabai = cabai.paginate(page=page, per_page=per_page, error_out=False)
    tomat = tomat.paginate(page=page, per_page=per_page, error_out=False)
    
    # Menggunakan list comprehension untuk menyederhanakan perhitungan total panen
    total_panen = [total.jml_panen for total in pangan]

    allDataCabai = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Cabai').order_by(desc(DataPangan.tanggal_bibit)).all()
    allDataTomat = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Tomat').order_by(desc(DataPangan.tanggal_bibit)).all()

    # Menggunakan list comprehension untuk menyederhanakan pembuatan list data statistik
    stat_cabai = [panenCabai.jml_panen for panenCabai in allDataCabai]
    tgl_panen_cabai = [panenCabai.tanggal_panen for panenCabai in allDataCabai]
    stat_tomat = [panenTomat.jml_panen for panenTomat in allDataTomat]
    tgl_panen_tomat = [panenTomat.tanggal_panen for panenTomat in allDataTomat]

    # Helper function to calculate percentage increase
    def calc_increase(data):
        if len(data) < 2 or 0 in data:
            return 0
        return round(((data[-1] - data[-2])/data[-2])*100)

    total_of_panen = sum(total_panen)
    totalPanenCabai = sum(stat_cabai)
    totalPanenTomat = sum(stat_tomat)

    if request.method == 'POST':
        kebun = request.form['kebun']
        komoditas = request.form['komoditas']
        jumlahBibit = request.form['jumlahBibit']
        tglBibit = request.form['tglBibit']
        tanggal_panen = datetime.today()
        estimasi_panen = tanggal_panen + timedelta(days=120)
        formatted_estimate = estimasi_panen.strftime('%Y-%m-%d')
        formatted_date = tanggal_panen.strftime('%Y-%m-%d')

        add_data = DataPangan(kebun=kebun, komoditas=komoditas,
                                tanggal_bibit=tglBibit, jml_bibit=jumlahBibit,
                                status='Penanaman', jml_panen=0, tanggal_panen=formatted_date, estimasi_panen=formatted_estimate,
                                user_id=current_user.id, kebun_id=current_user.kebun_id)
        db.session.add(add_data)
        db.session.commit()
        flash('Berhasil menginput data!', 'success')
        return redirect(request.referrer)

    return render_template('dashboard/data-pangan.html', 
                            max=max, min=min, 
                            allDataCabai=allDataCabai, allDataTomat=allDataTomat, 
                            kebun=kebun, user_data=user_data, 
                            kenaikan_cabai=calc_increase(stat_cabai), 
                            kenaikan_tomat=calc_increase(stat_tomat), 
                            stat_cabai=stat_cabai, stat_tomat=stat_tomat, 
                            cabai=cabai, tomat=tomat, pangan=pangan, 
                            total_panen=total_of_panen, 
                            totalPanenCabai=totalPanenCabai, 
                            totalPanenTomat=totalPanenTomat, 
                            tgl_panen_cabai=json.dumps(tgl_panen_cabai), 
                            tgl_panen_tomat=json.dumps(tgl_panen_tomat),
                            sort_by=sort_by, sort_order=sort_order)

@views.route('/dashboard/data-pangan/import', methods=['GET', 'POST'])
@login_required
def import_data_pangan():
    if current_user.role == 'admin':
        return redirect(url_for('admin_page.index'))
    
    from openpyxl import load_workbook

    if request.method == 'POST':
        import_type = request.form['import_type']
        excel_file = request.files['excel_file'] 

        if 'excel_file' not in request.files:
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)

        if excel_file.filename == '':
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)

        # Validasi ekstensi dan nama file
        if excel_file and report_allowed_file(excel_file.filename):
            if not allowed_report_stat(excel_file.filename, import_type):
                flash('Nama file harus sesuai format ("panen" atau "penanaman") dan sesuai dengan pilihan status produksi!', 'warning')
                return redirect(request.url)

            filename = secure_filename(excel_file.filename)
            
            try: 
                wb = load_workbook(excel_file)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2):
                    kebun = row[0].value
                    komoditas = row[1].value
                    jml_bibit = row[2].value
                    tanggal_bibit = row[3].value.strftime('%Y-%m-%d') if isinstance(row[3].value, datetime) else row[3].value 

                    # Menyederhanakan pembuatan objek DataPangan berdasarkan tipe impor
                    if import_type == 'penanaman':
                        data_pangan = DataPangan(kebun=kebun, komoditas=komoditas, 
                                                jml_bibit=jml_bibit, tanggal_bibit=tanggal_bibit, 
                                                status='Penanaman', jml_panen=0, tanggal_panen=0, 
                                                user_id=current_user.id, kelurahan_id=current_user.kelurahan_id)
                    elif import_type == 'panen':
                        jml_panen = row[4].value 
                        tanggal_panen = row[5].value.strftime('%Y-%m-%d') if isinstance(row[5].value, datetime) else row[5].value

                        data_pangan = DataPangan(kebun=kebun, komoditas=komoditas, 
                                                jml_bibit=jml_bibit, tanggal_bibit=tanggal_bibit, 
                                                status='Panen', jml_panen=jml_panen, tanggal_panen=tanggal_panen, 
                                                user_id=current_user.id, kelurahan_id=current_user.kelurahan_id)
                    else:
                        flash('Tipe impor tidak valid!', 'error')
                        return redirect(url_for('views.import_data_pangan'))
                    db.session.add(data_pangan)
        
                db.session.commit()
                flash('Data berhasil diimpor!', 'success')
                return redirect(url_for('views.dataproduksi'))
            except Exception as e:
                flash(f'Terjadi kesalahan saat memproses file: {e}', 'error')
                return redirect(request.url)
        else:
            flash('Ekstensi file tidak diizinkan. Unggah file Excel (.xlsx)!', 'error')
            return redirect(request.url) 

    return render_template('dashboard/import_data.html')

@views.route('/dashboard/data-pangan/delete_selected', methods=['POST'])
@login_required
def delete_selected_data_pangan():
    delete_ids = request.form.getlist('delete_ids')  # Get list of selected IDs

    if delete_ids:
        DataPangan.query.filter(DataPangan.id.in_(delete_ids)).delete(synchronize_session=False)
        db.session.commit()
        flash('Data yang dipilih berhasil dihapus!', 'warning')
    else:
        flash('Tidak ada data yang dipilih!', 'warning')

    return redirect(url_for('views.dataproduksi'))

@views.route('/dashboard/data-pangan/update-data/<int:id>', methods=['POST', 'GET'])
@login_required
def updatepangan(id):
    pangan = DataPangan.query.get_or_404(id)
    kel = Kebun.query.filter_by(id=current_user.kebun_id).first()

    updateProd = request.form.get('updateProduksi')

    if updateProd == 'updateProduksi':
        if request.method == 'POST':
            # Memperbarui data pangan dengan data dari form
            pangan.kebun = request.form['updateKebun']
            pangan.komoditas = request.form['updateKomoditas']
            pangan.jml_bibit = request.form['updateJumlahBibit']
            pangan.tanggal_bibit = request.form['updateTglBibit']

            db.session.commit()
            flash("Data berhasil diupdate!", "info")
            return redirect(url_for('views.dataproduksi'))
    elif updateProd == 'dataPanen':
        if request.method == 'POST':
            # Memperbarui data pangan dan kelurahan dengan data dari form
            pangan.status = 'Panen'
            pangan.jml_panen = request.form['updateJumlahPanen']
            pangan.tanggal_panen = request.form['updateTglPanen']
            pangan.kebun_id = kel.id

            kel.jml_panen = request.form['updateJumlahPanen']

            db.session.commit()
            flash("Data berhasil diupdate!", "info")
            return redirect(request.referrer)

@views.route('/dashboard/data-pangan/delete-data/<int:id>', methods=['GET'])
@login_required
def delete_data_pangan(id):
    data = DataPangan.query.get_or_404(id)
    db.session.delete(data)
    db.session.commit()
    return redirect(url_for('views.dataproduksi'))

# todo ============== PROFILE PAGE ==============
@views.route('/dashboard/profil', methods=['GET', 'POST'])
@login_required
def profil():
    if current_user.role == 'admin':
        return redirect(url_for('admin_page.index'))
    elif current_user.role == 'personal':
        return redirect(url_for('views.personalprofile'))

    user = User.query.filter_by(id=current_user.id).first()
    kebun = Kebun.query.filter_by(user_id=current_user.id).all()

    # Fetch regency, district, and village data based on stored IDs
    # regency_data = get_region_by_id('regencies', user.kota)
    # district_data = get_region_by_id('districts', user.kec)
    # village_data = get_region_by_id('villages', user.kelurahan)

    return render_template(
        'dashboard/profil.html',
        user=user,
        kebun=kebun,
        # regency_data=regency_data,
        # district_data=district_data,
        # village_data=village_data
    )

@views.route('/api/proxy/<path:url>')
def proxy(url):
    try:
        response = requests.get(f'https://emsifa.github.io/api-wilayah-indonesia/api/{url}')
        response.raise_for_status()
        return jsonify(response.json())
    except requests.exceptions.RequestException as e:
        print(f"Error in proxy: {str(e)}")  # Log error
        return jsonify({"error": str(e)}), 500

@views.route('/dashboard/profil/<int:id>/update', methods=['GET', 'POST'])
@login_required
def updateprofil(id):
    user = User.query.get_or_404(id)
    kebun = Kebun.query.filter_by(user_id=current_user.id).first()

    form_type = request.form.get('formType')

    if request.method == 'POST':
        if form_type == 'Data User':
            # Memperbarui data user dengan data dari form
            user.nama_lengkap = request.form['nama']
            user.username = request.form['username']
            user.pekerjaan = request.form['pekerjaan']
            user.kelamin = request.form['kelamin']
            
            # Hanya update data wilayah jika ada perubahan
            user.kota = request.form['regency'] if request.form['regency'] else user.kota
            user.kec = request.form['district'] if request.form['district'] else user.kec
            user.kelurahan = request.form['village'] if request.form['village'] else user.kelurahan

            user.bio = request.form['bio']

            db.session.commit()
            flash('Profil Berhasil Diubah', 'success')
            return redirect(url_for('views.profil')) 

        elif form_type == 'Data Kebun':
            if kebun:
                # Memperbarui data kelurahan dengan data dari form
                kebun.nama = request.form['nama_kebun']
                kebun.luas_kebun = request.form['luaskebun']
                kebun.koordinat = request.form['updateKoordinat']
            else:
                # Hanya buat kebun baru jika pengguna belum memiliki kebun
                new_kebun = Kebun(
                    user_id=user.id,
                    nama=request.form['nama_kebun'],
                    luas_kebun=request.form['luaskebun'],
                    koordinat=request.form['updateKoordinat']
                )
                db.session.add(new_kebun)

            db.session.commit()
            flash('Data Kebun Berhasil diubah!', 'success')
            return redirect(url_for('views.profil')) 

        else:
            flash('Tipe form tidak valid!', 'error')
            return redirect(url_for('views.profil'))

    return render_template('dashboard/profil.html', user=user, kebun=kebun)

@views.route('/personal/profile/update_picture/<int:id>', methods=['POST', 'GET'])
@login_required
def update_profile_picture(id):
    upload_folder = current_app.config['UPLOAD_FOLDER']
    user = User.query.get_or_404(id)
    if request.method == 'POST':
        if 'profile_pic' not in request.files:
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)
        file = request.files['profile_pic']
        if file.filename == '':
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)
        if file and picture_allowed_file(file.filename):
            filename = secrets.token_hex(8) + '_' + secure_filename(file.filename)
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)

            # Hapus foto profil lama jika ada
            if user.profile_pic:
                old_file_path = os.path.join(upload_folder, user.profile_pic)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            user.profile_pic = filename
            db.session.commit()
            flash('Foto profil berhasil diubah!', 'success')
        else:
            flash('File yang diizinkan hanya JPG, JPEG, dan PNG.', 'error')
    return redirect(url_for('views.profil'))

@views.route('/dashboard/profil/add_kebun', methods=['GET', 'POST'])
@login_required
def addkebun():
    user = User.query.get_or_404(current_user.id)
    form_type = request.form.get('formType')
    if request.method == 'POST':
        if form_type == 'Data Kebun':
            # Menambahkan data kebun baru
            nama = request.form['nama_kebun']
            luas = request.form['luaskebun']
            koordinat = request.form['koordinat']
            unique_id = generate_unique_id()
            if Kebun.query.filter_by(nama=nama).first():
                flash('Nama kebun sudah ada, silakan coba gunakan nama lain', 'warning')
                return redirect(url_for('views.profil'))
            else:
                add_kebun = Kebun(user_id=user.id, nama=nama, luas_kebun=luas, koordinat=koordinat, unique_id=unique_id)
                db.session.add(add_kebun)
                db.session.commit()
                flash('Kebun Berhasil Ditambahkan!', 'success')
                return redirect(url_for('views.profil'))
            
@views.route('/dashboard/profil/import_kebun', methods=['GET', 'POST'])
@login_required
def importkebun():
    from openpyxl import load_workbook

    if request.method == 'POST':
        excel_file = request.files['excel_file'] 

        if 'excel_file' not in request.files:
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)

        if excel_file.filename == '':
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)

        # Validasi ekstensi dan nama file
        if excel_file and report_allowed_file(excel_file.filename):
            # if not allowed_report_stat(excel_file.filename):
            #     flash('Nama file harus sesuai format template!', 'warning')
            #     return redirect(request.url)

            # filename = secure_filename(excel_file.filename)
            
            try: 
                wb = load_workbook(excel_file)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2):
                    nama_kebun = row[0].value
                    latitude = row[1].value
                    longitude = row[2].value
                    luas_kebun = row[3].value

                    if Kebun.query.filter_by(nama=nama_kebun).all:
                        flash('Kebun ini sudah ada, silakan gunakan nama lain', 'warning')
                        return redirect(url_for('views.profil'))
                    else:
                        # Menyederhanakan pembuatan objek DataPangan berdasarkan tipe impor
                        data_pangan = Kebun(nama=nama_kebun, koordinat=longitude+", "+latitude, luas_kebun=luas_kebun, user_id=current_user.id, unique_id=generate_unique_id())
                        db.session.add(data_pangan)

                db.session.commit()
                flash('Data berhasil diimpor!', 'success')
                return redirect(url_for('views.profil'))
            except Exception as e:
                flash(f'Terjadi kesalahan saat memproses file: {e}', 'error')
                return redirect(request.url)
        else:
            flash('Ekstensi file tidak diizinkan. Unggah file Excel (.xlsx)!', 'error')
            return redirect(request.url)

@views.route('/dashboard/profil/delete_kebun/<int:id>', methods=['GET', 'POST'])
@login_required
def delkebun(id):
    kebun = Kebun.query.get_or_404(id)
    
    # Memutuskan hubungan dengan user
    users = User.query.filter_by(kebun_id=kebun.id).all()
    for user in users:
        user.kebun_id = None
    
    # Hapus kebun
    db.session.delete(kebun)
    db.session.commit()
    
    flash('Kebun Berhasil Dihapus!', 'warning')
    return redirect(url_for('views.profil'))

@views.route('/dashboard/pengaturan', methods=['GET', 'POST'])
@login_required
def settings():
    if current_user.role == 'admin':
        return redirect(url_for('admin_page.index'))

    user = User.query.filter_by(id=current_user.id).first()
    return render_template('dashboard/settings.html', user=user)

@views.route('/dashboard/pengaturan/<int:id>/update-email', methods=['GET', 'POST'])
@login_required
def updateemail(id):
    user = User.query.get_or_404(id)
    password = request.form['userPass']

    if request.method == 'POST':
        if check_password_hash(current_user.password, password):
            user.email = request.form['email']
            db.session.commit()
            flash('Email berhasil diubah!', category='success')
            return redirect((request.referrer))
        else:
            flash('Kata sandi salah, silakan coba lagi!', category='error')
            return redirect(url_for('views.settings')) 

@views.route('/dashboard/pengaturan/<int:id>/update-password', methods=['GET', 'POST'])
@login_required
def updatepassword(id):
    user = User.query.get_or_404(id)
    oldpass = request.form['old-pass']
    newpass = request.form['new-pass']
    confnewpass = request.form['new-pass-conf']

    if request.method == 'POST':
        if check_password_hash(current_user.password, oldpass):
            if confnewpass != newpass:
                flash('Konfirmasi kata sandi tidak cocok!', 'error')
            else:
                user.password = generate_password_hash(newpass, method='pbkdf2')
                db.session.commit()
                flash('Kata sandi berhasil diperbarui!', category='success')
                return redirect((request.referrer)) 
        else:
            flash('Kata sandi salah, silakan coba lagi!', category='error')
            return redirect(url_for('views.settings'))

@views.route('/prakiraan-cuaca', methods=['GET', 'POST'])
def weather():
    return render_template('features/weather.html')

@views.route('/terms-and-conditions')
def terms():
    return render_template('terms_conditions.html')

# ========================= KELURAHAN SECTION =========================
@views.route('/peta-sebaran')
def mapbase():
    return render_template('kelurahan/map.html')

@views.route('/kelurahan-kulaba')
def kelkulaba():
    return render_template('kelurahan/kulaba.html')

@views.route('/kelurahan-sasa')
def kelsasa():
    return render_template('kelurahan/sasa.html')

@views.route('/kelurahan-kalumpang')
def kelkalumpang():
    return render_template('kelurahan/kalumpang.html')

@views.route('/kelurahan-santiong')
def kelsantiong():
    return render_template('kelurahan/santiong.html')

@views.route('/kelurahan-foramadiahi')
def kelforamadiahi():
    return render_template('kelurahan/foramadiahi.html')

@views.route('/kelurahan-tubo')
def keltubo():
    return render_template('kelurahan/tubo.html')

@views.route('/kelurahan-fitu')
def kelfitu():
    return render_template('kelurahan/fitu.html')

@views.route('/rindangtalk', methods=['GET', 'POST'])
def rindangtalk():
    questions = Forum.query.filter_by(created_by=current_user.id).all()
    fetch_ahli_email = User.query.filter_by(role='ahli').all()
    ahli_emails = fetch_ahli_email

    print(ahli_emails)

    if request.method == 'POST':
        nama = request.form['nama_lengkap']
        email = request.form['email']
        question = request.form['question']

        try:
            add_question = Forum(question=question, created_by=current_user.id)
            db.session.add(add_question)
            db.session.commit()
            forum_email(user_email=email, question=question)
            flash('Pertanyaan anda telah terkirim', 'success')
            return redirect(request.referrer)
        except:
            flash('Terjadi kesalahan saat mengirimkan pertanyaan, silakan coba kembali', 'error')
            return redirect(request.referrer)
    return render_template('features/rindangtalk.html', questions=questions)

@views.route('/avatar/<string:name>', methods=['GET', 'POST'])
def get_avatar(name):
    # Encoding untuk menghindari masalah karakter khusus
    encoded_name = name.encode('utf-8').decode('ascii', 'ignore')
    # Membangun URL avatar dengan aman
    avatar_url = f"https://ui-avatars.com/api/?name={encoded_name}&background=random&length=1&bold=true&rounded=true"
    return redirect(avatar_url) # Redirect ke URL avatar

def forum_email(user_email, question):
    try:
        subject = "Pertanyaan Terkirim ke RindangTalk"
        msg = Message(subject=subject, sender=('official@rindang.net'), recipients=[user_email], body=f'Anda mengirim pertanyaan ke ahli dengan dengan detail sebagai berikut: {question}')
        
        with mail.connect() as conn:
            conn.send(msg)
        current_app.logger.info(f"Email sent successfully to {user_email}")
    except smtplib.SMTPAuthenticationError:
        current_app.logger.error("SMTP Authentication Error. Please check your username and password.")
        raise
    except smtplib.SMTPException as e:
        current_app.logger.error(f"SMTP error occurred: {str(e)}")
        raise
    except Exception as e:
        current_app.logger.error(f"Failed to send email: {str(e)}")
        current_app.logger.exception("Email sending error")
        raise

def forum_email_to_ahli(user_email, user_name, question):
    try:
        subject = "Anda "
        msg = Message(subject=subject, sender=('official@rindang.net'), recipients=[user_email], body=f'Anda mendapatkan pertanyaan dari seorang pengguna dengan nama {user_name}, yaitu: {question}')
        
        with mail.connect() as conn:
            conn.send(msg)
        current_app.logger.info(f"Email sent successfully to {user_email}")
    except smtplib.SMTPAuthenticationError:
        current_app.logger.error("SMTP Authentication Error. Please check your username and password.")
        raise
    except smtplib.SMTPException as e:
        current_app.logger.error(f"SMTP error occurred: {str(e)}")
        raise
    except Exception as e:
        current_app.logger.error(f"Failed to send email: {str(e)}")
        current_app.logger.exception("Email sending error")
        raise

@views.route('/rindangtalk/update_question/<int:id>', methods=['GET', 'POST'])
@login_required
def update_question(id):
    data = Forum.query.get_or_404(id)
    return render_template('features/update_question.html', data=data)

@views.route('/rindangtalk/delete_question/<int:id>', methods=['GET', 'POST'])
@login_required
def delete_question(id):
    question = Forum.query.get_or_404(id)
    db.session.delete(question)
    db.session.commit()
    flash('Berhasil menghapus pertanyaan!', 'warning')
    return redirect(url_for('views.personal'))

@views.route('/rindangpedia')
def rindangpedia():
    articles = Artikel.query.all()
    return render_template('features/rindangpedia.html', articles=articles, shorten=shorten)

class MyHomeView(AdminIndexView):
    @expose('/')
    def index(self):
        user = User.query.all()
        return self.render('admin/index.html', user=user)

@views.route('/sitemap.xml')
def site_map():
    articles = sorted(flatpages, key=lambda item:item.meta['published'], reverse=False)
    return render_template('sitemap.xml', articles=articles, base_url="https://rindang.net")

@views.route('/robots.txt')
def robots_txt():
    return send_from_directory(current_app.static_folder, 'robots.txt')

admin = Admin(index_view=MyHomeView())